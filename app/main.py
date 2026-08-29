import os

from contextlib import asynccontextmanager
from pathlib import Path
import unicodedata
from fastapi import Depends, File, FastAPI, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from app.security import SESSION_SECRET, is_valid_password

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

import app.models
from app.database import Base, DATABASE_URL, engine, get_db
from app.models import Category, ImportBatch, Product, StockMovement
from app.schemas import (
    CategoryCreate,
    CategoryUpdate,
    CategoryResponse,
    ProductCreate,
    ProductResponse,
    ProductUpdate,
    StockMovementCreate,
    StockMovementResponse,
)


@asynccontextmanager
async def lifespan(app: FastAPI):
    # No SQLite local, mantém a criação automática usada durante o desenvolvimento.
    # No PostgreSQL da nuvem, as tabelas serão criadas somente pelo Alembic.
    if DATABASE_URL.startswith("sqlite"):
        Base.metadata.create_all(bind=engine)

    yield


APP_DIR = Path(__file__).resolve().parent

app = FastAPI(
    title="Inventory API",
    version="0.1.0",
    lifespan=lifespan,
)

COOKIE_SECURE = os.getenv("COOKIE_SECURE", "false").lower() == "true"

app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    same_site="lax",
    https_only=COOKIE_SECURE, 
)

app.mount("/static", StaticFiles(directory=APP_DIR / "static"), name="static")
templates = Jinja2Templates(directory=APP_DIR / "templates")


@app.get("/", include_in_schema=False)
def home():
    return RedirectResponse(url="/login", status_code=303)


@app.get("/health")
def health_check(db: Session = Depends(get_db)):
    return {
        "status": "ok",
        "database_backend": engine.url.get_backend_name(),
        "counts": {
            "categories": db.scalar(select(func.count(Category.id))),
            "products": db.scalar(select(func.count(Product.id))),
            "stock_movements": db.scalar(select(func.count(StockMovement.id))),
            },
        }


@app.post(
    "/categories",
    response_model=CategoryResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_category(
    category_data: CategoryCreate,
    db: Session = Depends(get_db),
):
    existing_category = db.scalar(
        select(Category).where(Category.name == category_data.name)
    )

    if existing_category:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Category already exists.",
        )

    category = Category(name=category_data.name)
    db.add(category)
    db.commit()
    db.refresh(category)

    return category



@app.get("/categories", response_model=list[CategoryResponse])
def list_categories(db: Session = Depends(get_db)):
    return db.scalars(
        select(Category).order_by(Category.name)
    ).all()

@app.put(
    "/categories/{category_id}",
    response_model=CategoryResponse,
)
def update_category(
    category_id: int,
    category_data: CategoryUpdate,
    db: Session = Depends(get_db),
):
    category = db.get(Category, category_id)

    if category is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Categoria não encontrada.",
        )

    normalized_name = category_data.name.strip()

    if len(normalized_name) < 2:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Informe um nome com pelo menos 2 caracteres.",
        )

    existing_category = db.scalar(
        select(Category).where(
            func.lower(Category.name) == normalized_name.lower(),
            Category.id != category_id,
        )
    )

    if existing_category:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Já existe uma categoria com este nome.",
        )

    category.name = normalized_name
    db.commit()
    db.refresh(category)

    return category



@app.post(
    "/products",
    response_model=ProductResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_product(
    product_data: ProductCreate,
    db: Session = Depends(get_db),
):
    category = db.get(Category, product_data.category_id) 

    if category is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Category not found.",
        )


    product = Product(**product_data.model_dump())
    db.add(product)
    db.commit()
    db.refresh(product)

    return product


@app.get("/products", response_model=list[ProductResponse])
def list_products(
    category_id: int | None = None,
    db: Session = Depends(get_db),
):
    statement = select(Product).order_by(Product.name)

    if category_id is not None:
        statement = statement.where(Product.category_id == category_id)

    return db.scalars(statement).all()

@app.get("/products/low-stock", response_model=list[ProductResponse])
def list_low_stock_products(db: Session = Depends(get_db)):
    return db.scalars(
        select(Product)
        .where(Product.quantity <= Product.minimum_quantity)
        .order_by(Product.quantity, Product.name)
    ).all()


@app.post(
    "/stock-movements",
    response_model=StockMovementResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_stock_movement(
    movement_data: StockMovementCreate,
    db: Session = Depends(get_db),
):
    product = db.get(Product, movement_data.product_id)

    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found.",
        )

    if (
        movement_data.movement_type == "out"
        and movement_data.quantity > product.quantity
    ):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Insufficient stock for this movement.",
        )

    if movement_data.movement_type == "in":
        product.quantity += movement_data.quantity
    else:
        product.quantity -= movement_data.quantity

    movement = StockMovement(**movement_data.model_dump())
    db.add(movement)
    db.commit()
    db.refresh(movement)

    return movement

@app.get(
    "/products/{product_id}/movements",
    response_model=list[StockMovementResponse],
)
def list_product_movements(
    product_id: int,
    db: Session = Depends(get_db),
):
    product = db.get(Product, product_id)

    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Product not found.",
        )

    return db.scalars(
        select(StockMovement)
        .where(StockMovement.product_id == product_id)
        .order_by(StockMovement.created_at.desc())
    ).all()


def redirect_to_login() -> RedirectResponse:
    return RedirectResponse(url="/login", status_code=303)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if request.session.get("warehouse_access"):
        return RedirectResponse(url="/warehouse", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="login.html",
        context={"error": None},
    )

@app.post("/login", response_class=HTMLResponse)
def login(request: Request, password: str = Form(...)):
    if not is_valid_password(password):
        return templates.TemplateResponse(
            request=request,
            name="login.html",
            context={"error": "Senha incorreta. Tente novamente."},
            status_code=401,
        )

    request.session["warehouse_access"] = True
    return RedirectResponse(url="/warehouse", status_code=303)


@app.post("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


@app.get("/warehouse", response_class=HTMLResponse)
def warehouse_page(
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    products = db.scalars(
        select(Product)
        .options(selectinload(Product.category))
        .order_by(Product.name)
    ).all()

    low_stock_products = [
        product
        for product in products
        if product.quantity <= product.minimum_quantity
    ]

    categories = db.scalars(
        select(Category)
        .where(Category.parent_id.is_(None))
        .order_by(Category.name)
    ).all()

    notice = request.session.pop("notice", None)

    return templates.TemplateResponse(
        request=request,
        name="warehouse.html",
        context={
            "products": products,
            "categories": categories,
            "low_stock_products": low_stock_products,
            "notice": notice,
        },
    )

@app.put("/products/{product_id}", response_model=ProductResponse)
def update_product(
    product_id: int,
    product_data: ProductUpdate,
    db: Session = Depends(get_db),
):
    product = db.get(Product, product_id)

    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Produto não encontrado.",
        )

    category = db.get(Category, product_data.category_id)

    if category is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Categoria não existente.",
        )

    product.name = product_data.name.strip()
    product.category_id = product_data.category_id
    product.minimum_quantity = product_data.minimum_quantity

    db.commit()
    db.refresh(product)

    return product


@app.delete("/products/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
):
    product = db.get(Product, product_id)

    if product is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Produto não encontrado.",
        )

    movements = db.scalars(
        select(StockMovement).where(
            StockMovement.product_id == product.id
        )
    ).all()

    for movement in movements:
        db.delete(movement)

    db.delete(product)
    db.commit()

@app.get('/warehouse/movement/{movement_type}', response_class=HTMLResponse)
def movement_form(
    movement_type: str,
    request: Request,
    db: Session = Depends(get_db),
):

    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    if movement_type not in {"in", "out"}:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)

    products = db.scalars(select(Product).order_by(Product.name)).all()

    return templates.TemplateResponse(
        request=request,
        name="movement_form.html",
        context={
            "products": products,
            "movement_type": movement_type,
        },
    )


@app.post("/warehouse/movement/{movement_type}")
def submit_movement(
    movement_type: str,
    request: Request,
    product_id: int = Form(...),
    quantity: int = Form(...),
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    if movement_type not in {"in", "out"}:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)

    if quantity <= 0:
        request.session["notice"] = "A quantidade deve ser maior que zero."
        return RedirectResponse(
            url=f"/warehouse/movement/{movement_type}",
            status_code=303,
        )

    product = db.get(Product, product_id)
    if product is None:
        request.session["notice"] = "Item não encontrado."
        return RedirectResponse(
            url=f"/warehouse/movement/{movement_type}",
            status_code=303,
        )

    if movement_type == "out" and quantity > product.quantity:
        request.session["notice"] = (
            f"Não há saldo suficiente de {product.name} para essa saída."
        )
        return RedirectResponse(
            url=f"/warehouse/movement/{movement_type}",
            status_code=303,
        )

    if movement_type == "in":
        product.quantity += quantity
    else:
        product.quantity -= quantity

    db.add(
        StockMovement(
            product_id=product.id,
            movement_type=movement_type,
            quantity=quantity,
        )
    )
    db.commit()

    movement_label = "entrada" if movement_type == "in" else "saída"
    request.session["notice"] = (
        f"{movement_label.capitalize()} de {quantity} unidade(s) "
        f"de {product.name} registrada com sucesso."
    )
    return RedirectResponse(url="/warehouse", status_code=303)



@app.post("/warehouse/products/new", response_class=HTMLResponse)
def submit_new_product(
    request: Request,
    name: str = Form(...),
    initial_quantity: int = Form(0),
    minimum_quantity: int = Form(3),
    category_id: int = Form(...),
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    values = {
        "name": name,
        "initial_quantity": initial_quantity,
        "minimum_quantity": minimum_quantity,
        "category_id": category_id,
    }

    normalized_name = name.strip()

    if len(normalized_name) < 2:
        return render_new_product_form(
            request,
            db,
            error="Informe um nome com pelo menos 2 caracteres.",
            values=values,
            status_code=422,
        )

    if initial_quantity < 0 or minimum_quantity < 0:
        return render_new_product_form(
            request,
            db,
            error="As quantidades não podem ser negativas.",
            values=values,
            status_code=422,
        )

    category = db.get(Category, category_id)

    if category is None:
        return render_new_product_form(
            request,
            db,
            error="Selecione uma categoria válida.",
            values=values,
            status_code=404,
        )

    product = Product(
        name=normalized_name,
        quantity=0,
        minimum_quantity=minimum_quantity,
        category_id=category_id,
    )
    db.add(product)
    db.flush()

    if initial_quantity > 0:
        db.add(
            StockMovement(
                product_id=product.id,
                movement_type="in",
                quantity=initial_quantity,
            )
        )
        product.quantity = initial_quantity

    db.commit()

    request.session["notice"] = f"Item {product.name} cadastrado com sucesso."
    return RedirectResponse(url="/warehouse", status_code=303)


@app.get("/warehouse/categories/new", response_class=HTMLResponse)
def new_category_form(
    request: Request,
    parent_id: int | None = None,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    parent_category = db.get(Category, parent_id) if parent_id else None
    if parent_id and parent_category is None:
        request.session["category_notice"] = "Categoria principal não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(url="/warehouse/categories", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="category_form.html",
        context={"error": None, "parent_category": parent_category},
    )


@app.post("/warehouse/categories/new", response_class=HTMLResponse)
def submit_new_category(
    request: Request,
    name: str = Form(...),
    parent_id: int | None = Form(None),
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    normalized_name = name.strip()

    if len(normalized_name) < 2:
        return templates.TemplateResponse(
            request=request,
            name="category_form.html",
            context={"error": "Informe um nome com pelo menos 2 caracteres."},
            status_code=422,
        )

    existing_category = db.scalar(
        select(Category).where(Category.name == normalized_name)
    )

    if existing_category:
        return templates.TemplateResponse(
            request=request,
            name="category_form.html",
            context={"error": "Essa categoria já está cadastrada."},
            status_code=409,
        )

    # Verifica a categoria principal, quando estiver sendo criada uma subcategoria
    parent_category = db.get(Category, parent_id) if parent_id else None

    if parent_id and parent_category is None:
        return templates.TemplateResponse(
            request=request,
            name="category_form.html",
            context={
                "error": "A categoria principal não foi encontrada.",
                "parent_category": None,
            },
            status_code=422,
        )

    category = Category(
        name=normalized_name,
        parent_id=parent_id,
    )
    db.add(category)
    db.commit()

    request.session["notice"] = f"Categoria {category.name} cadastrada com sucesso."
    return RedirectResponse(url="/warehouse/categories", status_code=303)

@app.get("/warehouse/categories/{category_id}/edit", response_class=HTMLResponse)
def edit_category_form(
    category_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    category = db.get(Category, category_id)

    if category is None:
        request.session["category_notice"] = "Categoria não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(url="/warehouse/categories", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="category_edit_form.html",
        context={"category": category, "error": None, "values": {}},
    )


@app.post("/warehouse/categories/{category_id}/edit", response_class=HTMLResponse)
def submit_edit_category(
    category_id: int,
    request: Request,
    name: str = Form(...),
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    category = db.get(Category, category_id)

    if category is None:
        request.session["category_notice"] = "Categoria não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(url="/warehouse/categories", status_code=303)

    normalized_name = name.strip()

    if len(normalized_name) < 2:
        return templates.TemplateResponse(
            request=request,
            name="category_edit_form.html",
            context={
                "category": category,
                "error": "Informe um nome com pelo menos 2 caracteres.",
                "values": {"name": name},
            },
            status_code=422,
        )

    existing_category = db.scalar(
        select(Category).where(
            func.lower(Category.name) == normalized_name.lower(),
            Category.id != category_id,
        )
    )

    if existing_category:
        return templates.TemplateResponse(
            request=request,
            name="category_edit_form.html",
            context={
                "category": category,
                "error": "Já existe uma categoria com este nome.",
                "values": {"name": name},
            },
            status_code=409,
        )

    category.name = normalized_name
    db.commit()

    request.session["category_notice"] = (
        f"Categoria {category.name} atualizada com sucesso."
    )
    return RedirectResponse(url="/warehouse/categories", status_code=303)

@app.get("/warehouse/categories/{category_id}", response_class=HTMLResponse)
def category_products_page(
    category_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    category = db.get(Category, category_id)

    if category is None:
        request.session["category_notice"] = "Categoria não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(url="/warehouse/categories", status_code=303)

    products = db.scalars(
        select(Product)
        .where(Product.category_id == category_id)
        .order_by(Product.name)
    ).all()

    import_batches = db.scalars(
        select(ImportBatch)
        .where(ImportBatch.category_id == category_id)
        .order_by(ImportBatch.created_at.desc())
    ).all()

    return templates.TemplateResponse(
        request=request,
        name="category_products.html",
        context={
            "category": category,
            "products": products,
            "import_batches": import_batches,
        },
    )

@app.get("/warehouse/categories/{category_id}/import", response_class=HTMLResponse)
def category_import_page(
    category_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    category = db.get(Category, category_id)

    if category is None:
        request.session["category_notice"] = "Categoria não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(
            url="/warehouse/categories",
            status_code=303,
        )

    return templates.TemplateResponse(
        request=request,
        name="category_import.html",
        context={
            "category": category,
            "error": None,
            "imported_count": None,
        },
    )


@app.post("/warehouse/categories/{category_id}/import", response_class=HTMLResponse )
def submit_category_import(
    category_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    category = db.get(Category, category_id)

    if category is None:
        request.session["category_notice"] = "Categoria não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(
            url="/warehouse/categories",
            status_code=303,
        )

    filename = (file.filename or "").lower()

    if not filename.endswith((".xlsx", ".xlsm")):
        return templates.TemplateResponse(
            request=request,
            name="category_import.html",
            context={
                "category": category,
                "error": "Envie um arquivo Excel .xlsx ou .xlsm.",
                "imported_count": None,
            },
            status_code=422,
        )

    try:
        workbook = load_workbook(
            file.file,
            read_only=True,
            data_only=True,
        )
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ValueError("A planilha está vazia.")

        def normalize_header(value):
            text = str(value or "").strip().casefold()
            text = unicodedata.normalize("NFKD", text)
            text = "".join(
                character
                for character in text
                if not unicodedata.combining(character)
            )
            return "".join(
                character
                for character in text
                if character.isalnum()
            )

        headers = [normalize_header(value) for value in rows[0]]

        name_aliases = {
            "nome",
            "name",
            "produto",
            "item",
            "material",
            "descricao",
            "description",
            "artigo",
            "mercadoria",
        }
        quantity_aliases = {
            "quantidade",
            "qtd",
            "qty",
            "estoque",
            "saldo",
            "quantidadeatual",
            "estoqueatual",
            "saldoatual",
        }
        minimum_aliases = {
            "estoqueminimo",
            "quantidademinima",
            "minimo",
            "min",
            "reposicao",
            "alertareposicao",
        }

        def find_exact_index(aliases, excluded=None):
            excluded = excluded or set()
            for index, header in enumerate(headers):
                if index not in excluded and header in aliases:
                    return index
            return None

        name_index = find_exact_index(name_aliases)
        minimum_index = find_exact_index(minimum_aliases)
        quantity_index = find_exact_index(
            quantity_aliases,
            excluded={minimum_index} if minimum_index is not None else set(),
        )

        if name_index is None:
            for index, header in enumerate(headers):
                if index != minimum_index and header:
                    name_index = index
                    break

        if name_index is None:
            raise ValueError(
                "Não foi possível identificar a coluna do nome do item."
            )

        if quantity_index is None:
            for index, header in enumerate(headers):
                if index in {name_index, minimum_index} or not header:
                    continue

                values_to_test = [
                    row[index]
                    for row in rows[1:]
                    if index < len(row)
                    and row[index] is not None
                    and str(row[index]).strip()
                ]

                if any(
                    isinstance(value, (int, float))
                    or str(value).strip().replace(",", ".").replace(".", "", 1).isdigit()
                    for value in values_to_test
                ):
                    quantity_index = index
                    break

        products_to_add = []
        seen_names = set()

        for line_number, row in enumerate(rows[1:], start=2):
            if not any(
                value is not None and str(value).strip()
                for value in row
            ):
                continue

            name = str(row[name_index] or "").strip()
            if not name:
                raise ValueError(
                    f"Linha {line_number}: informe o nome do item."
                )

            normalized_name = name.casefold()
            if normalized_name in seen_names:
                raise ValueError(
                    f"Linha {line_number}: o item '{name}' "
                    "aparece mais de uma vez na planilha."
                )

            already_exists = db.scalar(
                select(Product).where(
                    Product.category_id == category_id,
                    func.lower(Product.name) == normalized_name,
                )
            )

            if already_exists:
                raise ValueError(
                    f"Linha {line_number}: o item '{name}' "
                    "já existe nesta categoria."
                )

            try:
                raw_quantity = (
                    row[quantity_index]
                    if quantity_index is not None
                    and quantity_index < len(row)
                    else None
                )
                quantity = (
                    int(
                        float(
                            str(raw_quantity)
                            .strip()
                            .replace(",", ".")
                        )
                    )
                    if raw_quantity is not None
                    and str(raw_quantity).strip()
                    else 0
                )

                raw_minimum = (
                    row[minimum_index]
                    if minimum_index is not None
                    and minimum_index < len(row)
                    else None
                )
                minimum_quantity = (
                    int(
                        float(
                            str(raw_minimum)
                            .strip()
                            .replace(",", ".")
                        )
                    )
                    if raw_minimum is not None
                    and str(raw_minimum).strip()
                    else 3
                )

            except (TypeError, ValueError):
                raise ValueError(
                    f"Linha {line_number}: quantidade e "
                    "estoque_minimo devem ser números inteiros."
                )

            if quantity < 0:
                raise ValueError(
                    f"Linha {line_number}: quantidade não pode ser negativa."
                )

            if minimum_quantity < 0:
                raise ValueError(
                    f"Linha {line_number}: estoque_minimo não pode "
                    "ser negativo."
                )

            seen_names.add(normalized_name)
            products_to_add.append(
                {
                    "name": name,
                    "quantity": quantity,
                    "minimum_quantity": minimum_quantity,
                }
            )

        if not products_to_add:
            raise ValueError(
                "A planilha não contém linhas válidas para importar."
            )

        import_batch = ImportBatch(
            category_id=category_id,
            filename=file.filename or "importacao.xlsx",
        )
        db.add(import_batch)
        db.flush()

        for data in products_to_add:
            product = Product(
                name=data["name"],
                quantity=data["quantity"],
                minimum_quantity=data["minimum_quantity"],
                category_id=category_id,
                import_batch_id=import_batch.id,
            )
            db.add(product)
            db.flush()

            if data["quantity"] > 0:
                db.add(
                    StockMovement(
                        product_id=product.id,
                        movement_type="in",
                        quantity=data["quantity"],
                    )
                )

        db.commit()
        imported_count = len(products_to_add)


    except ValueError as exc:
        db.rollback()
        return templates.TemplateResponse(
            request=request,
            name="category_import.html",
            context={
                "category": category,
                "error": str(exc),
                "imported_count": None,
            },
            status_code=422,
        )

    except Exception:
        db.rollback()
        return templates.TemplateResponse(
            request=request,
            name="category_import.html",
            context={
                "category": category,
                "error": "Não foi possível ler a planilha. "
                "Confirme se o arquivo está íntegro e no formato Excel.",
                "imported_count": None,
            },
            status_code=422,
        )

    request.session["category_notice"] = (
        f"{imported_count} item(ns) importado(s) em {category.name}."
    )
    request.session["category_notice_kind"] = "success"
    return RedirectResponse(
        url=f"/warehouse/categories/{category_id}",
        status_code=303,
    )

@app.post("/warehouse/import-batches/{batch_id}/undo")
def undo_import_batch(
    batch_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    import_batch = db.get(ImportBatch, batch_id)

    if import_batch is None:
        request.session["category_notice"] = "Importação não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(
            url="/warehouse/categories",
            status_code=303,
        )

    category_id = import_batch.category_id

    products = db.scalars(
        select(Product).where(
            Product.import_batch_id == import_batch.id
        )
    ).all()

    for product in products:
        movements = db.scalars(
            select(StockMovement).where(
                StockMovement.product_id == product.id
            )
        ).all()

        for movement in movements:
            db.delete(movement)

        db.delete(product)

    filename = import_batch.filename
    db.delete(import_batch)
    db.commit()

    request.session["category_notice"] = (
        f"A importação {filename} foi desfeita com sucesso."
    )
    request.session["category_notice_kind"] = "success"
    return RedirectResponse(
        url=f"/warehouse/categories/{category_id}",
        status_code=303,
    )


@app.get("/warehouse/categories", response_class=HTMLResponse)
def categories_page(
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    root_categories = db.scalars(
        select(Category)
        .where(Category.parent_id.is_(None))
        .order_by(Category.name)
    ).all()

    categories = []

    for category in root_categories:
        product_count = db.scalar(
            select(func.count(Product.id)).where(
                Product.category_id == category.id
            )
        ) or 0

        subcategories = db.scalars(
            select(Category)
            .where(Category.parent_id == category.id)
            .order_by(Category.name)
        ).all()

        categories.append(
            {
                "id": category.id,
                "name": category.name,
                "product_count": product_count,
                "subcategories": subcategories,
            }
        )


    category_notice = request.session.pop("category_notice", None)
    notice_kind = request.session.pop("category_notice_kind", "success")

    return templates.TemplateResponse(
        request=request,
        name="categories.html",
        context={
            "categories": categories,
            "category_notice": category_notice,
            "notice_kind": notice_kind,
        },
    )

def render_new_product_form(
    request: Request,
    db: Session,
    error: str | None = None,
    values: dict | None = None,
    status_code: int = 200,
):
    categories = db.scalars(select(Category).order_by(Category.name)).all()

    return templates.TemplateResponse(
        request=request,
        name="product_form.html",
        context={
            "categories": categories,
            "error": error,
            "values": values or {},
        },
        status_code=status_code,
    )

def render_edit_product_form(
    request: Request,
    db: Session,
    product: Product,
    error: str | None = None,
    values: dict | None = None,
    status_code: int = 200,
):
    categories = db.scalars(select(Category).order_by(Category.name)).all()

    return templates.TemplateResponse(
        request=request,
        name="product_edit_form.html",
        context={
            "product": product,
            "categories": categories,
            "error": error,
            "values": values or {},
        },
        status_code=status_code,
    )


@app.get("/warehouse/products/{product_id}/edit", response_class=HTMLResponse)
def edit_product_form(
    product_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    product = db.get(Product, product_id)

    if product is None:
        request.session["notice"] = "Item não encontrado."
        return RedirectResponse(url="/warehouse", status_code=303)

    return render_edit_product_form(request, db, product)


@app.post("/warehouse/products/{product_id}/edit", response_class=HTMLResponse)
def submit_edit_product(
    product_id: int,
    request: Request,
    name: str = Form(...),
    category_id: int = Form(...),
    minimum_quantity: int = Form(...),
    quantity: int | None = Form(None),
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    product = db.get(Product, product_id)

    if product is None:
        request.session["notice"] = "Item não encontrado."
        return RedirectResponse(url="/warehouse", status_code=303)

    if quantity is None:
        quantity = product.quantity

    values = {
        "name": name,
        "category_id": category_id,
        "minimum_quantity": minimum_quantity,
        "quantity": quantity,
    }
    normalized_name = name.strip()

    if len(normalized_name) < 2:
        return render_edit_product_form(
            request,
            db,
            product,
            error="Informe um nome com pelo menos 2 caracteres.",
            values=values,
            status_code=422,
        )

    if minimum_quantity < 0:
        return render_edit_product_form(
            request,
            db,
            product,
            error="O estoque mínimo não pode ser negativo.",
            values=values,
            status_code=422,
        )

    if quantity < 0:
        return render_edit_product_form(
            request,
            db,
            product,
            error="A quantidade não pode ser negativa.",
            values=values,
            status_code=422,
    )

    category = db.get(Category, category_id)

    if category is None:
        return render_edit_product_form(
            request,
            db,
            product,
            error="Selecione uma categoria válida.",
            values=values,
            status_code=404,
        )

    previous_quantity = product.quantity
    quantity_difference = quantity - previous_quantity

    product.name = normalized_name
    product.category_id = category_id
    product.minimum_quantity = minimum_quantity
    product.quantity = quantity

    if quantity_difference != 0:
        db.add(
            StockMovement(
                product_id=product.id,
                movement_type="adjustment",
                quantity=quantity_difference,
            )
        )

    db.commit()

    request.session["notice"] = f"Item {product.name} atualizado com sucesso."
    return RedirectResponse(url="/warehouse", status_code=303)

@app.post("/warehouse/products/{product_id}/delete")
def submit_delete_product(
    product_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    product = db.get(Product, product_id)

    if product is None:
        request.session["notice"] = "Item não encontrado."
        return RedirectResponse(url="/warehouse", status_code=303)

    movements = db.scalars(
        select(StockMovement).where(
            StockMovement.product_id == product.id
        )
    ).all()

    for movement in movements:
        db.delete(movement)


    product_name = product.name
    db.delete(product)
    db.commit()

    request.session["notice"] = f"Item {product_name} removido com sucesso."
    return RedirectResponse(url="/warehouse", status_code=303)



@app.get("/warehouse/products/new", response_class=HTMLResponse)
def new_product_form(
    request: Request,
    category_id: int | None = None,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    values = {"category_id": category_id} if category_id else {}
    return render_new_product_form(request, db, values=values)



@app.post("/warehouse/products/new", response_class=HTMLResponse)
def submit_new_product(
    request: Request,
    name: str = Form(...),
    initial_quantity: int = Form(0),
    minimum_quantity: int = Form(3),
    category_id: int = Form(...),
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    values = {
        "name": name,
        "initial_quantity": initial_quantity,
        "minimum_quantity": minimum_quantity,
        "category_id": category_id,
    }

    normalized_name = name.strip()

    if len(normalized_name) < 2:
        return render_new_product_form(
            request,
            db,
            error="Informe um nome com pelo menos 2 caracteres.",
            values=values,
            status_code=422,
        )

    if initial_quantity < 0 or minimum_quantity < 0:
        return render_new_product_form(
            request,
            db,
            error="As quantidades não podem ser negativas.",
            values=values,
            status_code=422,
        )

    category = db.get(Category, category_id)

    if category is None:
        return render_new_product_form(
            request,
            db,
            error="Selecione uma categoria válida.",
            values=values,
            status_code=404,
        )

    product = Product(
        name=normalized_name,
        quantity=0,
        minimum_quantity=minimum_quantity,
        category_id=category_id,
    )
    db.add(product)
    db.flush()

    if initial_quantity > 0:
        db.add(
            StockMovement(
                product_id=product.id,
                movement_type="in",
                quantity=initial_quantity,
            )
        )
        product.quantity = initial_quantity

    db.commit()

    request.session["notice"] = f"Item {product.name} cadastrado com sucesso."
    return RedirectResponse(url="/warehouse", status_code=303)


@app.post("/warehouse/categories/{category_id}/delete")
def delete_category(
    category_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("warehouse_access"):
        return redirect_to_login()

    category = db.get(Category, category_id)

    if category is None:
        request.session["category_notice"] = "Categoria não encontrada."
        request.session["category_notice_kind"] = "warning"
        return RedirectResponse(url="/warehouse/categories", status_code=303)

    # Busca as subcategorias diretas da categoria principal.
    subcategories = db.scalars(
        select(Category).where(Category.parent_id == category.id)
    ).all()

    category_ids = [category.id] + [subcategory.id for subcategory in subcategories]

    # Busca os itens da categoria e de suas subcategorias.
    products = db.scalars(
        select(Product).where(Product.category_id.in_(category_ids))
    ).all()

    product_ids = [product.id for product in products]

    # Remove primeiro as movimentações para não violar a chave estrangeira.
    if product_ids:
        movements = db.scalars(
            select(StockMovement).where(
                StockMovement.product_id.in_(product_ids)
            )
        ).all()
        for movement in movements:
            db.delete(movement)

    # Depois remove os itens.
    for product in products:
        db.delete(product)

    # Depois remove as subcategorias.
    for subcategory in subcategories:
        db.delete(subcategory)

    # Por fim remove a categoria principal.
    category_name = category.name
    db.delete(category)
    db.commit()

    request.session["category_notice"] = (
        f"Categoria {category_name}, subcategorias, itens e movimentações "
        "foram removidos com sucesso."
    )
    request.session["category_notice_kind"] = "success"
    return RedirectResponse(url="/warehouse/categories", status_code=303)
